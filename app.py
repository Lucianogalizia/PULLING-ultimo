# =============================================================================
# app.py
# =============================================================================
from flask import Flask, request, redirect, url_for, render_template, flash, jsonify
import pandas as pd
import numpy as np
import datetime
import os
import re
import unicodedata
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from geopy.distance import geodesic
from difflib import SequenceMatcher


# Configuración de la aplicación Flask
app = Flask(__name__)
app.secret_key = "super_secret_key"
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Diccionario global para simular el "estado de sesión"
data_store = {}

# =============================================================================
# Funciones Auxiliares Generales
# =============================================================================
def normalize_text(text):
    if not isinstance(text, str):
        return text
    text = text.strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ASCII','ignore').decode('utf-8')
    return re.sub(r'\s+', ' ', text)

def extract_number(s):
    m = re.search(r'(\d+)', str(s))
    return m.group(1) if m else ""

def extract_letters(s):
    return "".join(re.findall(r'[A-Z]+', str(s).upper()))

def custom_normalize_pozo(user_pozo, coord_list, letter_threshold=0.5):
    if not isinstance(user_pozo, str):
        return user_pozo
    user_pozo = user_pozo.strip()
    num = extract_number(user_pozo)
    let = extract_letters(user_pozo)
    candidates = [c for c in coord_list if extract_number(c) == num and num]
    if not candidates:
        return user_pozo
    if len(candidates) == 1:
        return candidates[0]
    best, best_ratio = user_pozo, 0
    for cand in candidates:
        ratio = SequenceMatcher(None, let, extract_letters(cand)).ratio()
        if ratio > best_ratio:
            best_ratio, best = ratio, cand
    return best if best_ratio >= letter_threshold else user_pozo

# =============================================================================
# 1️⃣ Paso 1: Carga y marcado inicial
# =============================================================================
def step1_load_and_mark(file_path):
    wb = load_workbook(file_path, data_only=True)
    if "dataset" not in wb.sheetnames:
        raise ValueError("Hoja 'dataset' no encontrada.")
    ws = wb["dataset"]

    header = [cell.value or "" for cell in ws[1]]
    col_map = {i: normalize_text(header[i]) for i in range(len(header))}
    obs_idx    = next((i for i,cn in col_map.items() if "observac" in cn), None)
    pozo_idx   = next((i for i,cn in col_map.items() if "pozo"      in cn), None)
    equipo_idx = next((i for i,cn in col_map.items() if "equipo"    in cn), None)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {"CELESTE": False}
        # 1) Descarta rojos en OBSERVACIONES
        if obs_idx is not None:
            c = row[obs_idx]
            if (c.fill and c.fill.fgColor
               and c.fill.fgColor.type=="rgb"
               and c.fill.fgColor.rgb.upper()=="FFFF0000"):
                continue
        # 2) Filtra EQUIPO b2/b3
        if equipo_idx is not None:
            c = row[equipo_idx]
            if c.value and any(p in str(c.value).lower() for p in ("b2","b3")):
                continue
        # 3) Marca pozos celestes
        if pozo_idx is not None:
            c = row[pozo_idx]
            if (c.fill and c.fill.fgColor
               and c.fill.fgColor.type=="rgb"
               and c.fill.fgColor.rgb.upper()=="FF00FFFF"):
                row_data["CELESTE"] = True
        # 4) Recolecta valores
        for i, cell in enumerate(row):
            row_data[header[i]] = cell.value
        rows.append(row_data)

    df1 = pd.DataFrame(rows)
    preview1 = df1.head(20)
    return df1, preview1

# =============================================================================
# 2️⃣ Paso 2: Filtros básicos y limpieza
# =============================================================================
def step2_basic_filters(df1):
    df2 = df1.copy()
    if "Pérdida [m3/d]" in df2.columns:
        df2["Pérdida [m3/d]"] = pd.to_numeric(df2["Pérdida [m3/d]"], errors="coerce")
        df2.sort_values("Pérdida [m3/d]", ascending=False, inplace=True)
    if "Plan [Si/No]" in df2.columns:
        df2 = df2[df2["Plan [Si/No]"] == 1]

    critical = ["Activo","POZO","X","Y","Pérdida [m3/d]","Plan [Si/No]","Plan [Hs/INT]","EQUIPO"]
    for c in critical:
        if c in df2.columns:
            df2 = df2[df2[c].notnull() & (df2[c] != 0)]

    if "EQUIPO" in df2.columns:
        df2 = df2[df2["EQUIPO"].apply(lambda v: True 
            if not isinstance(v,str) 
            else not any(p in normalize_text(v) for p in ("fb","pesado","z inyector","z recupero")))]

    preview2 = df2.head(20)
    return df2, preview2

# =============================================================================
# 3️⃣ Paso 3: Normalización de POZO
# =============================================================================
def step3_normalize_pozos(df2):
    df3 = df2.copy()
    coords = pd.read_excel("coordenadas.xlsx", engine="openpyxl")
    coords["POZO_TMP"] = coords["POZO"].astype(str).str.strip().str.upper()
    coord_list = coords["POZO_TMP"].tolist()

    df3["POZO_TMP"] = df3["POZO"].astype(str).str.strip().str.upper()
    merged = df3.merge(coords[["POZO_TMP","POZO"]], on="POZO_TMP", how="left", suffixes=("","_coord"))
    merged["POZO"] = merged.apply(
        lambda r: r["POZO_coord"]
                  if pd.notnull(r["POZO_coord"])
                  else custom_normalize_pozo(r["POZO"], coord_list),
        axis=1
    )
    df3 = merged.drop(columns=["POZO_TMP","POZO_coord"])
    preview3 = df3.head(20)
    return df3, preview3

# =============================================================================
# 4️⃣ Paso 4: Merge de coordenadas
# =============================================================================
def step4_merge_coords(df3):
    coords = pd.read_excel("coordenadas.xlsx", engine="openpyxl")
    coords["GEO_LATITUDE"]  = coords["GEO_LATITUDE"].astype(str).str.replace(",",".").astype(float)
    coords["GEO_LONGITUDE"] = coords["GEO_LONGITUDE"].astype(str).str.replace(",",".").astype(float)

    df4 = df3.merge(coords[["POZO","GEO_LATITUDE","GEO_LONGITUDE"]], on="POZO", how="left")
    missing = df4[df4["GEO_LATITUDE"].isnull() | df4["GEO_LONGITUDE"].isnull()]["POZO"].unique().tolist()
    df4 = df4.dropna(subset=["GEO_LATITUDE","GEO_LONGITUDE"])
    return df4, missing

# =============================================================================
# 5️⃣ Paso 5: Finalizar y renombrar
# =============================================================================
def step5_finalize(df4):
    df5 = df4.rename(columns={
        "Activo":"ZONA",
        "Pérdida [m3/d]":"NETA [M3/D]",
        "Plan [Hs/INT]":"TIEMPO PLANIFICADO",
        "Batería":"BATERÍA"
    }).copy()
    df5["PROD_DT"] = datetime.date.today().strftime("%Y-%m-%d")
    df5["RUBRO"]   = "ESPERA DE TRACTOR"
    order = ["POZO","NETA [M3/D]","PROD_DT","RUBRO",
             "GEO_LATITUDE","GEO_LONGITUDE","BATERÍA","ZONA","TIEMPO PLANIFICADO"]
    return df5[[c for c in order if c in df5.columns]]

# =============================================================================
# Orquestador: reemplaza al antiguo process_excel
# =============================================================================
def process_excel(file_path):
    df1, preview1   = step1_load_and_mark(file_path)
    df2, _          = step2_basic_filters(df1)
    df3, _          = step3_normalize_pozos(df2)
    df4, missing    = step4_merge_coords(df3)
    df_final        = step5_finalize(df4)
    pozos_celestes  = df1[df1["CELESTE"]==True]["POZO"].unique().tolist()
    return df_final, preview1, pozos_celestes

# =============================================================================
# Rutas de la Aplicación Flask
# =============================================================================
@app.route("/")
def index():
    return redirect(url_for("upload_file"))


 from flask import Flask, request, redirect, url_for, render_template, flash

 import pandas as pd
 # … resto de imports …

 # (todas tus funciones step1_load_and_mark, step2_basic_filters, etc. y process_excel)

#
# ─── RUTAS AJAX PARA PROCESAMIENTO EN 5 PASOS ───────────────────────────
#
@app.route("/process/step1", methods=["POST"])
def process_step1():
    if "excel_file" not in request.files:
        return jsonify({"error": "No se encontró el archivo."}), 400
    f = request.files["excel_file"]
    filename = secure_filename(f.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(path)
    try:
        df1, preview1 = step1_load_and_mark(path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    data_store["df1"] = df1
    return jsonify(preview1.to_dict(orient="records"))

@app.route("/process/step2", methods=["POST"])
def process_step2():
    df1 = data_store.get("df1")
    df2, preview2 = step2_basic_filters(df1)
    data_store["df2"] = df2
    return jsonify(preview2.to_dict(orient="records"))

@app.route("/process/step3", methods=["POST"])
def process_step3():
    df2 = data_store.get("df2")
    df3, preview3 = step3_normalize_pozos(df2)
    data_store["df3"] = df3
    return jsonify(preview3.to_dict(orient="records"))

@app.route("/process/step4", methods=["POST"])
def process_step4():
    df3 = data_store.get("df3")
    df4, missing = step4_merge_coords(df3)
    data_store["df4"] = df4
    return jsonify({"missing_pozos": missing})

@app.route("/process/step5", methods=["POST"])
def process_step5():
    df4 = data_store.get("df4")
    df_final = step5_finalize(df4)
    data_store["df"] = df_final    # reemplaza el df limpio
    return df_final.to_html(classes="table table-striped", index=False)
# ────────────────────────────────────────────────────────────────────────

@app.route("/upload", methods=["GET","POST"])
def upload_file():
    if request.method == "POST":
        if "excel_file" not in request.files:
            flash("No se encontró el archivo en la solicitud.")
            return redirect(request.url)
        file = request.files["excel_file"]
        if file.filename == "":
            flash("No se seleccionó ningún archivo.")
            return redirect(request.url)
        filename = secure_filename(file.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(path)
        try:
            df_clean, preview_df, pozos_celestes = process_excel(path)
        except Exception as e:
            flash(f"Error al procesar el Excel: {e}")
            return redirect(request.url)
        data_store["df"]           = df_clean
        data_store["celeste_pozos"] = pozos_celestes
        flash("Archivo procesado exitosamente. A continuación, el preview.")
        return render_template("upload_success.html",
                               preview=preview_df.to_html(classes="table table-striped", index=False))
    return render_template("upload.html")

@app.route("/filter", methods=["GET","POST"])
def filter_zonas():
    if "df" not in data_store:
        flash("Debes subir un archivo Excel primero.")
        return redirect(url_for("upload_file"))
    df = data_store["df"]
    zonas = sorted(df["ZONA"].unique().tolist())
    if request.method == "POST":
        sel = request.form.getlist("zonas")
        if not sel:
            flash("Debes seleccionar al menos una zona.")
            return redirect(request.url)
        try:
            cnt = int(request.form.get("pulling_count", 3))
        except:
            cnt = 3
        df_f = df[df["ZONA"].isin(sel)].copy()
        data_store["df_filtrado"]       = df_f
        data_store["pozos_disponibles"] = sorted(df_f["POZO"].tolist())
        data_store["pulling_count"]     = cnt
        flash(f"Zonas: {', '.join(sel)} | Pullings: {cnt}")
        return redirect(url_for("select_pulling"))
    checkbox_html = "".join(f'<input type="checkbox" name="zonas" value="{z}"> {z}<br>' for z in zonas)
    return render_template("filter_zonas.html", checkbox_html=checkbox_html)

@app.route("/select_pulling", methods=["GET","POST"])
def select_pulling():
    if "df_filtrado" not in data_store:
        flash("Debes filtrar las zonas primero.")
        return redirect(url_for("filter_zonas"))
    df_f        = data_store["df_filtrado"]
    disponibles = data_store["pozos_disponibles"]
    cnt         = data_store["pulling_count"]
    celestes    = data_store["celeste_pozos"]
    if request.method == "POST":
        pulling_data = {}
        usados = []
        for i in range(1, cnt+1):
            p = request.form.get(f"pulling_pozo_{i}")
            pulling_data[f"Pulling {i}"] = {"pozo": p, "tiempo_restante": 0.0}
            usados.append(p)
        if len(usados) != len(set(usados)):
            flash("No puedes repetir pozos.")
            return redirect(request.url)
        data_store["pulling_data"]      = pulling_data
        data_store["pozos_disponibles"] = [p for p in disponibles if p not in usados]
        flash("Pullings confirmados.")
        return redirect(url_for("assign"))
    form_html = ""
    for i in range(1, cnt+1):
        default = celestes[i-1] if i-1 < len(celestes) else None
        opts = "".join(f'<option value="{p}"{" selected" if p==default else ""}>{p}</option>' for p in disponibles)
        form_html += f"<h4>Pulling {i}</h4><select name='pulling_pozo_{i}'>{opts}</select><hr>"
    return render_template("select_pulling.html", form_html=form_html)

@app.route("/assign", methods=["GET"])
def assign():
    if "pulling_data" not in data_store:
        flash("Completa Pulling primero.")
        return redirect(url_for("select_pulling"))
    df   = data_store["df"]
    pdat = data_store["pulling_data"]
    plst = list(pdat.items())
    occupied = set()

    def coef(ref, cand):
        r1 = df[df["POZO"]==ref].iloc[0]
        r2 = df[df["POZO"]==cand].iloc[0]
        d  = geodesic((r1["GEO_LATITUDE"],r1["GEO_LONGITUDE"]),
                      (r2["GEO_LATITUDE"],r2["GEO_LONGITUDE"])).kilometers
        n  = r2["NETA [M3/D]"]; t = r2["TIEMPO PLANIFICADO"]
        den = t + 0.5*d
        return (n/den if den else 0), d

    def asign(a):
        avail = data_store["pozos_disponibles"]
        for pull,_ in plst:
            ref = a[pull][-1][0] if a[pull] else pdat[pull]["pozo"]
            cands = [(p,*coef(ref,p)) for p in avail if p not in occupied]
            cands.sort(key=lambda x:(-x[1],x[2]))
            if cands:
                best = cands[0]
                a[pull].append(best)
                occupied.add(best[0])
        return a

    asigs = {p:[] for p,_ in plst}
    for _ in ("N+1","N+2","N+3"):
        asigs = asign(asigs)

    matrix = []
    for pull,_ in plst:
        po_act = pdat[pull]["pozo"]
        n_act  = df[df["POZO"]==po_act]["NETA [M3/D]"].iloc[0]
        tr     = pdat[pull]["tiempo_restante"]
        sels   = asigs[pull][:3] + [("N/A",0,0)]*3
        row    = [pull, po_act, n_act, tr]
        for p,cf,d in sels[:3]:
            n = df[df["POZO"]==p]["NETA [M3/D]"].iloc[0] if p!="N/A" else 0
            row += [p,n,cf,d]
        matrix.append(row)

    cols = ["Pulling","Pozo Actual","Neta Actual","Tiempo Restante (h)",
            "N+1","Neta N+1","Coef N+1","Dist N+1 (km)",
            "N+2","Neta N+2","Coef N+2","Dist N+2 (km)",
            "N+3","Neta N+3","Coef N+3","Dist N+3 (km)"]
    dfp = pd.DataFrame(matrix, columns=cols)
    table_html = dfp.style.hide_index().format(precision=2).render()
    flash("Asignación completada.")
    return render_template("assign_result.html", table=table_html)

if __name__ == "__main__":
    app.run(debug=True)




